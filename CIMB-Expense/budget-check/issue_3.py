import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from google.colab import files

data_json = [
  {
    "bank_account_number": "701094834900",
    "div_code": "1611-24008-1036",
    "div_name": "CIMB Niaga Conventional Cianjur - Cokroaminoto NDB - ONE RM",
    "email_initiator": "asep.yusuf@cimbniaga.co.id",
    "position": "3021970-Client Relationship Manager - (BsM 1) - Business Owners - CNJ",
    "ref_id": 162921
  },
  {
    "bank_account_number": "701094834900",
    "div_code": "1611-24008-1036",
    "div_name": "CIMB Niaga Conventional Cianjur - Cokroaminoto NDB - ONE RM",
    "email_initiator": "asep.yusuf@cimbniaga.co.id",
    "position": "3021970-Client Relationship Manager - (BsM 1) - Business Owners - CNJ",
    "ref_id": 162920
  },
  {
    "bank_account_number": "763183805800",
    "div_code": "1611-23032-1036",
    "div_name": "CIMB Niaga Conventional Cileungsi NDB - ONE RM",
    "email_initiator": "agus.hendriawan@cimbniaga.co.id",
    "position": "5395882-Client Relationship Manager - (BsM 1) - Business Owners - CLS",
    "ref_id": 162733
  },
  {
    "bank_account_number": "763988040900",
    "div_code": "1611-23020-1038",
    "div_name": "CIMB Niaga Conventional Bogor - Juanda NDB - CRM AFFLUENT",
    "email_initiator": "gawok.khatelu@cimbniaga.co.id",
    "position": "5421918-Client Relationship Manager - (BsM 1) - Affluent - BGR",
    "ref_id": 162729
  },
  {
    "bank_account_number": "762781559200",
    "div_code": "1611-23026-1036",
    "div_name": "CIMB Niaga Conventional Bogor - V Point NDB - ONE RM",
    "email_initiator": "rr.mastorani@cimbniaga.co.id",
    "position": "5373707-Client Relationship Manager - (BsM 1) - Business Owners - BVP",
    "ref_id": 162596
  },
  {
    "bank_account_number": "762776166700",
    "div_code": "1611-47077-2022",
    "div_name": "BRANCH SALES ROLLOUT-077",
    "email_initiator": "dedy.sutomo@cimbniaga.co.id",
    "position": "5372869-Contact Center Service Process Optimization",
    "ref_id": 162142
  },
  {
    "bank_account_number": "703738917800",
    "div_code": "1612-47077-0823",
    "div_name": "Syariah Consumer Financing Sales",
    "email_initiator": "sasha.anggea@cimbniaga.co.id",
    "position": "3032877-Syariah Mortgage Sales Management Specialist - BAA",
    "ref_id": 162098
  },
  {
    "bank_account_number": "701517262900",
    "div_code": "1611-00014-0102",
    "div_name": "Jakarta - Niaga Tower-COMMERCIAL HIGH END 2",
    "email_initiator": "savina.salim@cimbniaga.co.id",
    "position": "3025713-Sr Business Relationship Manager - ComBa - Jakarta 1 - Area II",
    "ref_id": 161958
  },
  {
    "bank_account_number": "701517262900",
    "div_code": "1611-00014-0102",
    "div_name": "Jakarta - Niaga Tower-COMMERCIAL HIGH END 2",
    "email_initiator": "savina.salim@cimbniaga.co.id",
    "position": "3025713-Sr Business Relationship Manager - ComBa - Jakarta 1 - Area II",
    "ref_id": 161957
  },
  {
    "bank_account_number": "705267014100",
    "div_code": "1611-47077-1129",
    "div_name": "1611-47077-1129 Mortgage Business - Strategic & Support - 1611-47077-1129",
    "email_initiator": "selfia.pratiwi@cimbniaga.co.id",
    "position": "3036230-Mortgage Business Manager - Jakarta East & Jatim 3",
    "ref_id": 161920
  },
  {
    "bank_account_number": "705411561000",
    "div_code": "1611-16033-1425",
    "div_name": "TMT-S&D CP JKT - TOMANG TOL-177",
    "email_initiator": "tri.annisa@cimbniaga.co.id",
    "position": "3036744-Private Wealth Relationship Manager - KC - TMT",
    "ref_id": 161884
  },
  {
    "bank_account_number": "762588467900",
    "div_code": "1611-29012-1036",
    "div_name": "CIMB Niaga Conventional Batam - Lumbung Rezeki NDB - ONE RM",
    "email_initiator": "desy.desy@cimbniaga.co.id",
    "position": "5348854-Client Relationship Manager - (BsM 2) - Business Owners - BTM",
    "ref_id": 161465
  },
  {
    "bank_account_number": "700218553200",
    "div_code": "1611-47077-7680",
    "div_name": "COBA OFFICE-COBA LOAN AGENCY",
    "email_initiator": "amalia.syafei@cimbniaga.co.id",
    "position": "5601348-Corporate Trust Sr Specialist",
    "ref_id": 161400
  },
  {
    "bank_account_number": "764104642900",
    "email_initiator": "yuniati.yuniati@cimbniaga.co.id",
    "position": "5600726-Preferred Relationship Manager 2 - (BsM1) - KCP - PHR",
    "ref_id": 161344
  },
  {
    "bank_account_number": "764104642900",
    "email_initiator": "yuniati.yuniati@cimbniaga.co.id",
    "position": "5600726-Preferred Relationship Manager 2 - (BsM1) - KCP - PHR",
    "ref_id": 161342
  },
  {
    "bank_account_number": "705848678600",
    "email_initiator": "liana.pappa@cimbniaga.co.id",
    "position": "3038728-Client Relationship Manager - (BsM 1) - Business Owners - GJY",
    "ref_id": 160980
  },
  {
    "bank_account_number": "763644292400",
    "div_code": "1611-47077-1802",
    "div_name": "1611-47077-1802 - Alternate Channel-077",
    "email_initiator": "astrid.putri@cimbniaga.co.id",
    "position": "5407520-Digital Channel Strategy, Experience & Governance Head",
    "ref_id": 160703
  },
  {
    "bank_account_number": "764207681000",
    "div_code": "1611-39027-1034",
    "div_name": "CONVENTIONAL  Surabaya - Kembang Jepun NDB - EMERGING BUSINESS BANKING NDB",
    "email_initiator": "felita.randy@cimbniaga.co.id",
    "position": "5601956-Network EBB Relationship Manager - KCP - KMA",
    "ref_id": 160475
  },
  {
    "bank_account_number": "764072180700",
    "div_code": "1612-27006-0752",
    "div_name": "1612-27006-0752 CIMB Niaga Syariah Palembang SYARIAH - NETWORK SERVICE & OPERATION",
    "email_initiator": "ilmelda.damaiyanti@cimbniaga.co.id",
    "position": "5600247-Syariah Preferred Relationship Manager - KCS - SSPL",
    "ref_id": 159801
  },
  {
    "bank_account_number": "764218125700",
    "email_initiator": "jesslyn.aprilia@cimbniaga.co.id",
    "position": "5601911-Funding Relationship Manager - Team 2 - Jakarta 2",
    "ref_id": 159547
  },
  {
    "bank_account_number": "764202088700",
    "div_code": "1611-34070-1036",
    "div_name": "CIMB Niaga Conventional-Solo - Manahan-NDB - ONE RM",
    "email_initiator": "yefta.saputra@cimbniaga.co.id",
    "position": "5601910-Client Relationship Manager - (BsM 1) - Business Owners - MNH",
    "ref_id": 159258
  },
  {
    "bank_account_number": "764072180700",
    "div_code": "1612-27006-0752",
    "div_name": "1612-27006-0752 CIMB Niaga Syariah Palembang SYARIAH - NETWORK SERVICE & OPERATION",
    "email_initiator": "ilmelda.damaiyanti@cimbniaga.co.id",
    "position": "5600247-Syariah Preferred Relationship Manager - KCS - SSPL",
    "ref_id": 158883
  },
  {
    "bank_account_number": "701813284300",
    "div_code": "1611-34070-1036",
    "div_name": "CIMB Niaga Conventional-Solo - Manahan-NDB - ONE RM",
    "email_initiator": "frieda.christiana@cimbniaga.co.id",
    "position": "3004291-Client Relationship Manager - (BsM 2) - Business Owners - MNH - 2",
    "ref_id": 158605
  },
  {
    "bank_account_number": "763395703200",
    "div_code": "1611-47077-6710",
    "div_name": "HEAD OFFICE ANALYTICS",
    "email_initiator": "bani.fahlevi@cimbniaga.co.id",
    "position": "5402413-Sales Campaign Management Specialist",
    "ref_id": 158130
  },
  {
    "bank_account_number": "763988040900",
    "div_code": "1611-23020-1038",
    "div_name": "CIMB Niaga Conventional Bogor - Juanda NDB - CRM AFFLUENT",
    "email_initiator": "gawok.khatelu@cimbniaga.co.id",
    "position": "5421918-Client Relationship Manager - (BsM 1) - Affluent - BGR",
    "ref_id": 157167
  },
  {
    "bank_account_number": "763827903500",
    "div_code": "1611-23047-1038",
    "div_name": "CIMB Niaga Conventional KCP Bekasi - Summarecon NDB - CRM AFFLUENT",
    "email_initiator": "hanna.sartika@cimbniaga.co.id",
    "position": "5413668-Client Relationship Manager - (BsM 1) - Affluent - SMB",
    "ref_id": 156803
  }
]

HEADER_MAPPING = {
    "ref_id": "Ref ID",
    "bank_account_number": "Bank Account Number",
    "div_code": "Div Code",
    "div_name": "Div Name",
    "email_initiator": "Email Initiator",
    "position": "Posisi",
}

COLUMN_ORDER = [
    "ref_id",
    "bank_account_number",
    "div_code",
    "div_name",
    "position",
    "email_initiator",
]

def normalize_json_to_df(data):
    if isinstance(data, dict):
        possible_keys = ["data", "results", "items", "records", "rows"]
        found = False
        for key in possible_keys:
            if key in data and isinstance(data[key], list):
                data = data[key]
                found = True
                break
        if not found:
            data = [data]

    if not isinstance(data, list):
        raise ValueError("Format JSON tidak dikenali. Pastikan data berupa list of records atau dict.")

    df = pd.json_normalize(data)

    existing_cols = list(df.columns)
    ordered_cols = [c for c in COLUMN_ORDER if c in existing_cols]
    remaining_cols = [c for c in existing_cols if c not in ordered_cols]
    df = df[ordered_cols + remaining_cols]

    new_headers = {}
    for col in df.columns:
        if col in HEADER_MAPPING:
            new_headers[col] = HEADER_MAPPING[col]
        else:
            new_headers[col] = col.replace("_", " ").title()
    df = df.rename(columns=new_headers)

    return df

df = normalize_json_to_df(data_json)

print(f"Total baris data: {len(df)}")
print(f"Total kolom: {len(df.columns)}")
df.head()

output_filename = "after_patching.xlsx"

def export_to_excel(df, filename, sheet_name="Data"):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            if pd.isna(value):
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.freeze_panes = "A2"

    for col_idx, col_name in enumerate(df.columns, start=1):
        max_length = len(str(col_name))
        for value in df.iloc[:, col_idx - 1]:
            if pd.notna(value):
                max_length = max(max_length, len(str(value)))
        adjusted_width = min(max_length + 4, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    ws.row_dimensions[1].height = 30

    wb.save(filename)
    print(f"File berhasil dibuat: {filename}")

export_to_excel(df, output_filename)

files.download(output_filename)